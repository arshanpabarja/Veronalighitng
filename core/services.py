import copy
from pathlib import Path
import jdatetime
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent

TEMPLATE_PATH = (
    BASE_DIR /
    "templates" /
    "factor_template.xlsx"
)


def to_jalali(date_obj):
    return jdatetime.date.fromgregorian(
        date=date_obj
    ).strftime("%Y/%m/%d")

def copy_row_style(ws, source_row, target_row):

    for col in range(1, ws.max_column + 1):

        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)

        if source.has_style:
            target._style = copy.copy(source._style)

        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)

        if source.number_format:
            target.number_format = source.number_format


def copy_merged_ranges(ws, source_row, target_row):

    merged_ranges = list(ws.merged_cells.ranges)

    for merged in merged_ranges:

        if (
            merged.min_row == source_row
            and
            merged.max_row == source_row
        ):

            offset = target_row - source_row

            try:
                ws.merge_cells(
                    start_row=merged.min_row + offset,
                    start_column=merged.min_col,
                    end_row=merged.max_row + offset,
                    end_column=merged.max_col
                )
            except Exception:
                pass


def add_extra_rows_for_items(ws, items_count):

    DEFAULT_ROWS = 5

    if items_count <= DEFAULT_ROWS:
        return

    extra_rows = items_count - DEFAULT_ROWS

    ws.insert_rows(30, amount=extra_rows)

    for row in range(30, 30 + extra_rows):

        copy_row_style(ws, 29, row)

        copy_merged_ranges(
            ws,
            29,
            row
        )


def fill_customer_data(ws, invoice):

    ws["BC1"] = invoice.serial

    ws["BC2"] = to_jalali(invoice.date)

    if invoice.expire_date:
        ws["BC3"] = to_jalali(
            invoice.expire_date
        )

    ws["G18"] = invoice.customer_name
    ws["I20"] = invoice.province
    ws["Q20"] = invoice.city
    ws["E22"] = invoice.address
    ws["AY22"] = invoice.customer_phone


def fill_items_and_calculate(ws, items):

    start_row = 25

    total_before_discount = 0
    total_discount = 0

    for index, item in enumerate(items, start=1):

        row = start_row + index - 1

        quantity = float(item.quantity)
        unit_price = int(item.unit_price)
        discount = int(item.discount)

        row_total = quantity * unit_price
        after_discount = row_total - discount
        tax = round(after_discount * 1)
        final_amount = after_discount + tax

        total_before_discount += row_total
        total_discount += discount

        # شماره ردیف
        ws[f"B{row}"] = index

        # نام کالا
        ws[f"D{row}"] = item.invoice_name

        # شرح
        ws[f"G{row}"] = item.description

        # تعداد
        ws[f"T{row}"] = quantity

        # واحد
        ws[f"W{row}"] = item.unit

        # مبلغ واحد
        ws[f"Z{row}"] = unit_price

        # مبلغ کل
        ws[f"AE{row}"] = row_total

        # تخفیف
        ws[f"AL{row}"] = discount

        # مبلغ بعد تخفیف
        ws[f"AQ{row}"] = after_discount

        # مالیات
        ws[f"AX{row}"] = tax

        # مبلغ نهایی
        ws[f"BC{row}"] = final_amount

    total_after_discount = (
        total_before_discount -
        total_discount
    )

    total_tax = round(
        total_after_discount * 1
    )

    total_final = (
        total_after_discount +
        total_tax
    )

    total_row = 30 + max(0, len(items) - 5)

    # جمع مبلغ کل
    ws[f"AE{total_row}"] = total_before_discount

    # جمع تخفیف
    ws[f"AL{total_row}"] = total_discount

    # جمع بعد از تخفیف
    ws[f"AQ{total_row}"] = total_after_discount

    # جمع مالیات
    ws[f"AX{total_row}"] = total_tax

    # جمع نهایی
    ws[f"BC{total_row}"] = total_final


def generate_invoice_excel(
    invoice,
    output_file
):

    wb = load_workbook(
        TEMPLATE_PATH
    )

    ws = wb.active

    fill_customer_data(
        ws,
        invoice
    )

    items = list(
        invoice.items.all()
    )

    add_extra_rows_for_items(
        ws,
        len(items)
    )

    fill_items_and_calculate(
        ws,
        items
    )

    wb.save(output_file)

    return output_file